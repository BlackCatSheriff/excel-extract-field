import os
import sys

from openpyxl import Workbook
from openpyxl import load_workbook


def init_ext_index_excle(sheet, ext_fields):
    display_dic = dict(zip(ext_fields, [0] * len(ext_fields)))
    # 处理提取字段在原始表中的序号， 初始化 display_dic
    for i, cell in enumerate(sheet['1']):
        if str(cell.value).upper() in ext_fields:
            display_dic[str(cell.value).upper()] = i
    return list(display_dic.values())


def do_row(row, ext_index):
    res = []
    for i in ext_index:
        res.append(row[i].value)
    return res


def save_file(save_path, data, fields):
    # 增量添加,如果文件存在就打开追加
    wb = Workbook()
    sheet = wb.active
    if os.path.exists(save_path):
        wb = load_workbook(save_path)
        sheet = wb[wb.sheetnames[0]]
    else:
        sheet.append(fields)
    for d in data:
        sheet.append(d)
    wb.save(save_path)
    wb.close()


def doit(args):
    open_file, save_path, ext_fields = parser(args)
    wb = load_workbook(open_file)
    sheets = wb.sheetnames
    # 目前就一份文件保存一下把，如果数据太多就在优化这里
    res = []
    #for si in range(len(sheets)):
    sheet = wb[sheets[0]]
    ext_index = init_ext_index_excle(sheet, ext_fields)
    first = False
    for row in sheet.rows:
        if first:
            res.append(do_row(row, ext_index))
        first = True

    wb.close()
    save_file(save_path, res, ext_fields)


def parser(args):
    extract_fields = []
    file_path = args[1]
    save_path = args[2]
    for i in range(3, len(args)):
        extract_fields.append(args[i].upper())  #全部到大写，，处理大小写不敏感
    return file_path, save_path, extract_fields


if __name__ == "__main__":
    # 命令行参数 提取文件路径 保存文件路径 args....
    try:
        doit(sys.argv)
        print("DONE")
    except:
        print("FAIL")
