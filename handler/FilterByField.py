import os
import sys

from openpyxl import Workbook
from openpyxl import load_workbook


def init_ext_index_excel(sheet, ext_field):
    # 处理提取字段在原始表中的序号
    for i, cell in enumerate(sheet['1']):
        if str(cell.value).upper() == ext_field:
            return i


def get_heads(sheet):
    res = []
    for cell in sheet['1']:
        res.append(str(cell.value))
    return res


def cells_to_list(cells, length):
    res = []
    for i in range(length):
        res.append(cells[i].value)
    return res


def save_file(save_path, data, sheetheads):
    # 增量添加,如果文件存在就打开追加
    wb = Workbook()
    sheet = wb.active
    if os.path.exists(save_path):
        wb = load_workbook(save_path)
        sheet = wb[wb.sheetnames[0]]
    else:
        sheet.append(sheetheads)
    for d in data:
        sheet.append(d)
    wb.save(save_path)
    wb.close()


def doit(args):
    open_file, ext_fields = parser(args)
    wb = load_workbook(open_file)
    sheets = wb.sheetnames
    # 目前就一份文件保存一下把，如果数据太多就在优化这里
    res = {}    # 数据按照 field 不同值生成字典，{'aa':[],'bb':[]}
    sheet = wb[sheets[0]]
    ext_index = init_ext_index_excel(sheet, ext_fields)
    sheetheads = get_heads(sheet)
    open_file_name = open_file.split('.')[0]
    first = False
    for i, row in enumerate(sheet.rows):
        if first:
            m_key = row[ext_index].value
            if res.get(m_key, -1) == -1:
                res[m_key] = list()
            res[m_key].append(i)
        first = True

    for K in res.keys():
        out_table = []
        for r in res.get(K):
            out_table.append(cells_to_list(sheet[str(r+1)], len(sheetheads)))
        save_file("{}_{}.xlsx".format(open_file_name, K), out_table, sheetheads)

    wb.close()


def parser(args):
    return args[1], args[2].upper()


if __name__ == "__main__":
    # 命令行参数 提取文件路径 保存文件路径 args....
    print(sys.argv)
    try:
        doit(sys.argv)
        print("DONE")
    except:
        print("FAIL")
